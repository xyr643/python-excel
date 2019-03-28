import openpyxl

def print_book_lists_excel(book_lists, book_tag_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    for i in range(len(book_tag_lists)):
        # ws = wb.active
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode()))  # utf8->unicode
    for i in range(len(book_tag_lists)):
        # 列名
        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版社'])
        count = 1
        for bl in book_lists[i]:
            # append按行写入Excel
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1
    save_path = 'book_list'
    for i in range(len(book_tag_lists)):
        save_path += ('-' + book_tag_lists[i].decode())
    save_path += '.xlsx'
    wb.save(save_path)
    
    if __name__ == '__main__':
      book_tag_lists = ['计算机','机器学习','linux','数据库','互联网']
      # book_lists为多个数组
      print_book_lists_excel(book_lists, book_tag_lists)
